import streamlit as st
from groq import Groq
import pytesseract
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import chromadb
from sentence_transformers import SentenceTransformer
import io
import os

# ── Page Config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="QA Test Generator",
    page_icon="🧪",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── CSS ────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600;700&family=IBM+Plex+Sans:wght@300;400;500;600&display=swap');

*, html, body, [class*="css"] {
    font-family: 'IBM Plex Sans', sans-serif;
}

.stApp {
    background: #F7F6F3;
}

/* Header */
.app-header {
    background: #1A1A2E;
    border-radius: 0 0 24px 24px;
    padding: 2.5rem 2rem 2rem;
    margin: -1rem -1rem 2rem -1rem;
    position: relative;
    overflow: hidden;
}
.app-header::after {
    content: '🧪';
    position: absolute;
    right: 2rem;
    top: 50%;
    transform: translateY(-50%);
    font-size: 5rem;
    opacity: 0.08;
}
.app-title {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 2rem;
    font-weight: 700;
    color: #E8F4FD;
    margin: 0;
    letter-spacing: -0.5px;
}
.app-sub {
    color: #7B8FA1;
    font-size: 0.9rem;
    margin-top: 0.4rem;
    font-weight: 300;
}
.tag {
    display: inline-block;
    background: rgba(255,255,255,0.07);
    border: 1px solid rgba(255,255,255,0.12);
    color: #A8C8E8;
    padding: 0.15rem 0.65rem;
    border-radius: 4px;
    font-size: 0.7rem;
    font-family: 'IBM Plex Mono', monospace;
    margin: 0.3rem 0.2rem 0 0;
    letter-spacing: 0.5px;
}

/* Cards */
.card {
    background: #FFFFFF;
    border: 1px solid #E8E5DF;
    border-radius: 12px;
    padding: 1.5rem;
    margin-bottom: 1rem;
}
.card-label {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.65rem;
    color: #9B8EA0;
    text-transform: uppercase;
    letter-spacing: 2px;
    margin-bottom: 0.5rem;
}

/* Pipeline steps */
.pipeline {
    display: flex;
    gap: 0.4rem;
    margin: 1rem 0;
    flex-wrap: wrap;
}
.pipe-step {
    flex: 1;
    min-width: 80px;
    background: #F0EDE8;
    border: 1px solid #E0DDD8;
    border-radius: 8px;
    padding: 0.6rem 0.4rem;
    text-align: center;
    font-size: 0.7rem;
    color: #8A8580;
    font-family: 'IBM Plex Mono', monospace;
}
.pipe-step.done {
    background: #E8F5E9;
    border-color: #A5D6A7;
    color: #2E7D32;
}
.pipe-step.active {
    background: #E3F2FD;
    border-color: #90CAF9;
    color: #1565C0;
    font-weight: 600;
}

/* Result box */
.result-preview {
    background: #1A1A2E;
    border-radius: 8px;
    padding: 1rem 1.2rem;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.75rem;
    color: #A8C8E8;
    white-space: pre-wrap;
    max-height: 220px;
    overflow-y: auto;
    margin-top: 0.5rem;
}

/* Metric */
.metric-strip {
    display: flex;
    gap: 1rem;
    margin: 1rem 0;
}
.metric {
    flex: 1;
    background: #FFFFFF;
    border: 1px solid #E8E5DF;
    border-radius: 10px;
    padding: 1rem;
    text-align: center;
}
.metric-val {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1.8rem;
    font-weight: 700;
    color: #1A1A2E;
}
.metric-lbl {
    font-size: 0.7rem;
    color: #9B8EA0;
    text-transform: uppercase;
    letter-spacing: 1px;
    margin-top: 0.2rem;
}

/* Success */
.success-box {
    background: #E8F5E9;
    border: 1px solid #A5D6A7;
    border-left: 4px solid #2E7D32;
    border-radius: 8px;
    padding: 1.2rem 1.5rem;
    margin: 1rem 0;
    color: #1B5E20;
    font-weight: 500;
}

/* Buttons */
.stButton > button {
    background: #1A1A2E !important;
    color: #E8F4FD !important;
    border: none !important;
    border-radius: 8px !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 0.82rem !important;
    font-weight: 600 !important;
    letter-spacing: 1px !important;
    padding: 0.65rem 1.5rem !important;
    width: 100% !important;
    transition: all 0.2s !important;
}
.stButton > button:hover {
    background: #2D2D4E !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 16px rgba(26,26,46,0.25) !important;
}
.stButton > button:disabled {
    background: #C8C5BF !important;
    color: #8A8580 !important;
    cursor: not-allowed !important;
}

/* Inputs */
.stTextInput > div > input,
.stTextArea > div > textarea {
    background: #FFFFFF !important;
    border: 1px solid #D8D5CF !important;
    border-radius: 8px !important;
    font-family: 'IBM Plex Sans', sans-serif !important;
    color: #1A1A2E !important;
}
.stTextInput > div > input:focus,
.stTextArea > div > textarea:focus {
    border-color: #1A1A2E !important;
    box-shadow: 0 0 0 2px rgba(26,26,46,0.1) !important;
}

/* Sidebar */
[data-testid="stSidebar"] {
    background: #FFFFFF;
    border-right: 1px solid #E8E5DF;
}
.sidebar-section {
    background: #F7F6F3;
    border: 1px solid #E8E5DF;
    border-radius: 10px;
    padding: 1rem;
    margin-bottom: 1rem;
}
.sidebar-title {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.65rem;
    text-transform: uppercase;
    letter-spacing: 2px;
    color: #9B8EA0;
    margin-bottom: 0.6rem;
}

/* Tabs */
.stTabs [data-baseweb="tab"] {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.78rem;
}

/* Hide Streamlit chrome */
#MainMenu, footer, header { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

# ── Session State ──────────────────────────────────────────────────────────────
for k in ["extracted_text", "parsed_req", "similar_cases",
          "test_cases", "excel_bytes", "stage", "tc_count"]:
    if k not in st.session_state:
        st.session_state[k] = None
if st.session_state.stage is None:
    st.session_state.stage = 0


# ── Helpers ────────────────────────────────────────────────────────────────────
def call_groq(prompt: str, api_key: str) -> str:
    client = Groq(api_key=api_key)
    response = client.chat.completions.create(
        model="llama-3.1-8b-instant",
        messages=[{"role": "user", "content": prompt}],
        max_tokens=2000
    )
    return response.choices[0].message.content


def run_ocr(image: Image.Image) -> str:
    return pytesseract.image_to_string(
        image.convert("RGB"), config="--psm 6"
    ).strip()


@st.cache_resource
def get_rag():
    embed = SentenceTransformer("all-MiniLM-L6-v2")
    db    = chromadb.Client()
    col   = db.get_or_create_collection("past_test_cases")
    return embed, col


def load_excel_to_rag(excel_bytes: bytes):
    embed, col = get_rag()
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    headers = [
        str(c.value).strip().lower() if c.value else f"col{i}"
        for i, c in enumerate(ws[1])
    ]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        text = "\n".join(
            f"{headers[i]}: {v}"
            for i, v in enumerate(row)
            if v is not None and i < len(headers)
        )
        if text.strip():
            emb = embed.encode(text).tolist()
            try:
                col.add(documents=[text], embeddings=[emb], ids=[f"tc_{count}"])
                count += 1
            except Exception:
                pass
    return count


def rag_retrieve(query: str) -> str:
    embed, col = get_rag()
    if col.count() == 0:
        return "No past test cases loaded."
    emb     = embed.encode(query).tolist()
    results = col.query(query_embeddings=[emb], n_results=min(3, col.count()))
    docs    = results["documents"][0] if results["documents"] else []
    return "\n\n---\n\n".join(docs) if docs else "No similar cases found."


def build_excel(test_cases_text: str, feature_name: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    hfill  = PatternFill("solid", fgColor="1A1A2E")
    afill  = PatternFill("solid", fgColor="EBF3FB")
    hfont  = Font(bold=True, color="E8F4FD", name="Calibri", size=11)
    nfont  = Font(name="Calibri", size=10)
    ca     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    la     = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    thin   = Side(style="thin", color="D0CCC8")
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["TC ID","Title","Preconditions","Test Steps","Expected Result","Priority","Status"]
    widths  = [10, 28, 25, 40, 32, 12, 12]

    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell            = ws.cell(row=1, column=c, value=h)
        cell.fill       = hfill
        cell.font       = hfont
        cell.alignment  = ca
        cell.border     = bdr
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 30

    lines = [
        l.strip() for l in test_cases_text.split("\n")
        if "|" in l and "---" not in l and l.strip() != "|"
    ]

    for ri, line in enumerate(lines, 2):
        parts = [p.strip() for p in line.strip("|").split("|")]
        while len(parts) < 7:
            parts.append("")
        if not parts[6]:
            parts[6] = "Not Run"
        for ci, val in enumerate(parts[:7], 1):
            cell           = ws.cell(row=ri, column=ci, value=val)
            cell.font      = nfont
            cell.border    = bdr
            cell.alignment = ca if ci in [1, 6, 7] else la
            if ri % 2 == 0:
                cell.fill = afill
        ws.row_dimensions[ri].height = 45

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"], ws2["B1"] = "Feature", feature_name
    ws2["A2"], ws2["B2"] = "Total Test Cases", f"=COUNTA('Test Cases'!A2:A1000)"
    ws2["A3"], ws2["B3"] = "Generated By", "AI Test Case Generator"
    for c in ["A1","A2","A3"]:
        ws2[c].font = Font(bold=True, name="Calibri")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# LAYOUT
# ══════════════════════════════════════════════════════════════════════════════

# ── Header ─────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="app-header">
    <div class="app-title">QA Test Case Generator</div>
    <div class="app-sub">Upload feature images or paste a description → Get professional test cases in Excel</div>
    <div style="margin-top:0.8rem">
        <span class="tag">GROQ AI</span>
        <span class="tag">TESSERACT OCR</span>
        <span class="tag">RAG</span>
        <span class="tag">CHROMADB</span>
        <span class="tag">EXCEL EXPORT</span>
    </div>
</div>
""", unsafe_allow_html=True)

# ── Sidebar ─────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="sidebar-title">⚙️ Configuration</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-section">', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-title">🔑 Groq API Key</div>', unsafe_allow_html=True)
    api_key = st.text_input("", type="password",
                            placeholder="Paste your gsk_... key",
                            label_visibility="collapsed")
    st.caption("Free key → [console.groq.com](https://console.groq.com)")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-section">', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-title">📊 Past Test Cases (RAG)</div>', unsafe_allow_html=True)
    st.caption("Upload your existing test cases Excel file")
    past_excel = st.file_uploader("", type=["xlsx","xls"],
                                  label_visibility="collapsed")
    if past_excel:
        with st.spinner("Loading into RAG..."):
            count = load_excel_to_rag(past_excel.read())
        st.success(f"✅ {count} test cases loaded")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-section">', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-title">🎛️ Settings</div>', unsafe_allow_html=True)
    num_cases       = st.slider("Number of test cases", 5, 30, 10)
    include_neg     = st.toggle("Include negative tests", value=True)
    include_edge    = st.toggle("Include edge cases",     value=True)
    st.markdown('</div>', unsafe_allow_html=True)

    # Pipeline tracker
    st.markdown("---")
    st.markdown('<div class="sidebar-title">📋 Pipeline Status</div>', unsafe_allow_html=True)
    stage    = st.session_state.stage or 0
    steps    = ["OCR", "Parse", "RAG", "Generate", "Excel"]
    for i, s in enumerate(steps):
        icon = "✅" if i < stage else ("⏳" if i == stage else "○")
        color = "#2E7D32" if i < stage else ("#1565C0" if i == stage else "#9B8EA0")
        st.markdown(
            f'<div style="color:{color};font-size:0.85rem;padding:0.2rem 0;">'
            f'{icon} {s}</div>',
            unsafe_allow_html=True
        )

# ── Main ────────────────────────────────────────────────────────────────────────
col_left, col_right = st.columns([1, 1], gap="large")

with col_left:
    # ── Input Section ────────────────────────────────────────────────────────
    st.markdown('<div class="card-label">📤 Feature Input</div>', unsafe_allow_html=True)

    input_tab1, input_tab2 = st.tabs(["🖼️ Upload Images", "📝 Type / Paste Text"])

    with input_tab1:
        images = st.file_uploader(
            "Upload one or more feature images",
            type=["jpg","jpeg","png","webp","bmp"],
            accept_multiple_files=True,
            label_visibility="collapsed"
        )
        if images:
            for img in images:
                st.image(img, caption=img.name, use_container_width=True)

    with input_tab2:
        manual_text = st.text_area(
            "Paste your feature description here",
            height=200,
            placeholder="""Example:
Feature: Add Issue Pattern
The user can select a Rule Pattern from a dropdown.
Dropdown options: Contains, Equals, Starts With, Ends With
User clicks Save to save the pattern.
Validation: Pattern field cannot be empty.""",
            label_visibility="collapsed"
        )

    st.markdown("")

    # Validate inputs
    has_image  = bool(images)
    has_text   = bool(manual_text and manual_text.strip())
    has_input  = has_image or has_text
    can_run    = has_input and bool(api_key)

    if not api_key:
        st.caption("🔑 Add your Groq API key in the sidebar")
    elif not has_input:
        st.caption("⬆️ Upload an image or type a description above")

    # ── Generate Button ───────────────────────────────────────────────────────
    if st.button("⚡ GENERATE TEST CASES", disabled=not can_run):

        # ── Stage 1: OCR ──────────────────────────────────────────────────────
        st.session_state.stage = 1
        combined_text = ""

        if has_image:
            with st.spinner("🔍 Reading images with OCR..."):
                for img_file in images:
                    img  = Image.open(img_file)
                    text = run_ocr(img)
                    if text:
                        combined_text += f"\n\n[From {img_file.name}]\n{text}"

        if has_text:
            combined_text += f"\n\n[Manual Description]\n{manual_text.strip()}"

        combined_text = combined_text.strip()
        if not combined_text:
            st.error("❌ Could not extract any text. Please add a text description.")
            st.stop()

        st.session_state.extracted_text = combined_text

        # ── Stage 2: Parse Requirements ───────────────────────────────────────
        st.session_state.stage = 2
        with st.spinner("🧠 Parsing requirements with Groq AI..."):
            parse_prompt = f"""You are a senior business analyst.
Read the feature story below and extract structured requirements.

Return in exactly this format:
Feature Name: [name]
Description: [brief description]
UI Elements: [list all buttons, fields, dropdowns]
User Actions: [list what user can do]
Expected Behaviors: [list expected outcomes]
Validations: [list all validation rules]
Edge Cases: [list potential edge cases]

Feature Story:
{combined_text}
"""
            parsed = call_groq(parse_prompt, api_key)
            st.session_state.parsed_req = parsed

        # ── Stage 3: RAG ──────────────────────────────────────────────────────
        st.session_state.stage = 3
        with st.spinner("📚 Searching past test cases..."):
            similar = rag_retrieve(parsed)
            st.session_state.similar_cases = similar

        # ── Stage 4: Generate Test Cases ──────────────────────────────────────
        st.session_state.stage = 4
        with st.spinner(f"⚙️ Generating {num_cases} test cases..."):
            neg_inst  = "Include negative test cases (invalid inputs, errors)." if include_neg  else ""
            edge_inst = "Include edge cases (empty, boundary, special chars)."  if include_edge else ""

            gen_prompt = f"""You are a senior QA engineer.
Generate exactly {num_cases} test cases for this feature.

Requirements:
{parsed}

Reference style from past test cases:
{similar}

{neg_inst}
{edge_inst}

CRITICAL RULES:
- Return ONLY the table rows, no headings, no extra text, no markdown
- Each row must follow this EXACT pipe-delimited format:
| TC001 | Test title | Preconditions | Step1; Step2; Step3 | Expected result | High |
- Priority must be: High, Medium, or Low
- Generate exactly {num_cases} rows
"""
            test_cases = call_groq(gen_prompt, api_key)
            st.session_state.test_cases = test_cases

            # Count rows
            lines = [l for l in test_cases.split("\n")
                     if "|" in l and "---" not in l]
            st.session_state.tc_count = len(lines)

        # ── Stage 5: Excel ────────────────────────────────────────────────────
        st.session_state.stage = 5
        with st.spinner("📊 Building Excel file..."):
            feature_name = "feature"
            if images:
                feature_name = os.path.splitext(images[0].name)[0]
            excel = build_excel(test_cases, feature_name)
            st.session_state.excel_bytes = excel

        st.rerun()

with col_right:
    st.markdown('<div class="card-label">📋 Results</div>', unsafe_allow_html=True)

    if not st.session_state.test_cases:
        st.markdown("""
        <div style="background:#FFFFFF;border:1px solid #E8E5DF;border-radius:12px;
                    padding:4rem 2rem;text-align:center;color:#C8C5BF;">
            <div style="font-size:3rem">⏳</div>
            <div style="margin-top:0.8rem;font-size:0.9rem;color:#9B8EA0;">
                Results will appear here after generation
            </div>
        </div>
        """, unsafe_allow_html=True)

    else:
        # Success
        tc = st.session_state.tc_count or 0
        st.markdown(f"""
        <div class="success-box">
            ✅ &nbsp; <strong>{tc} test cases generated successfully!</strong>
        </div>
        """, unsafe_allow_html=True)

        # Metrics
        chars = len(st.session_state.extracted_text or "")
        _, col_db = get_rag()
        rag_count = col_db.count()

        st.markdown(f"""
        <div class="metric-strip">
            <div class="metric">
                <div class="metric-val">{tc}</div>
                <div class="metric-lbl">Test Cases</div>
            </div>
            <div class="metric">
                <div class="metric-val">{chars}</div>
                <div class="metric-lbl">Chars Read</div>
            </div>
            <div class="metric">
                <div class="metric-val">{rag_count}</div>
                <div class="metric-lbl">RAG Cases</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Tabs
        t1, t2, t3 = st.tabs(["📊 Test Cases", "🔍 OCR Output", "🧠 Requirements"])

        with t1:
            st.markdown(
                f'<div class="result-preview">{st.session_state.test_cases}</div>',
                unsafe_allow_html=True
            )

        with t2:
            st.markdown(
                f'<div class="result-preview">{st.session_state.extracted_text or "—"}</div>',
                unsafe_allow_html=True
            )

        with t3:
            st.markdown(
                f'<div class="result-preview">{st.session_state.parsed_req or "—"}</div>',
                unsafe_allow_html=True
            )

        # Download
        st.markdown("---")
        fname = "feature"
        if images:
            fname = os.path.splitext(images[0].name)[0]

        st.download_button(
            label="⬇️ DOWNLOAD EXCEL TEST CASES",
            data=st.session_state.excel_bytes,
            file_name=f"{fname}_test_cases.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        if st.button("🔄 Start Over"):
            for k in ["extracted_text","parsed_req","similar_cases",
                      "test_cases","excel_bytes","tc_count"]:
                st.session_state[k] = None
            st.session_state.stage = 0
            st.rerun()
